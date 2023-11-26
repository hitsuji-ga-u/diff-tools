import abc
from enum import Enum
import csv
from typing import Any
import openpyxl as px
from pprint  import pprint
from pathlib import Path

class SignalExistError(Exception):
    def __init__(self, message='     Sinals already exists.   '):
        self.message = message
        super().__init__(self.message)


class Judgement(Enum):
    NOJUDGEMENT = 0
    SAME = 1
    ADD = 2
    DEL = 3
    CHANGE = 4


 



class Attribute(object):

    attribute_group = [
        {'min', '最小値'}
    ]

    def __init__(self, name, value=''):
        self.name = name
        self.value = value

    def __repr__(self):
        return f'Attribute(\'{self.name}\', \'{self.value}\')'

    def __eq__(self, other):
        return self.value == other.value and self.same_name(other)

    def set(self, value):
        return Attribute(self.name, value)

    def same_name(self, attr):
        if self.name.lower() == attr.name.lower():
            return True

        for group in self.attribute_group:
            if not self.name in group:
                continue
            if attr.name in group:
                return True

        return False

class UndifinedAttribute(Attribute):

    def __init__(self, name):
        super().__init__('Undifined')

    def __bool__(self):
        return False

    def set(self, value):
        return Attribute(self.name, '-')


class Recode(object):

    def __init__(self, key_idx, attributes):
        self.key = attributes[key_idx]
        self.attributes = attributes


    def __repr__(self):
        return f'Recode({self.attributes.index(self.key)}, {self.attributes})'

    def get_value(self, attr):
        for my_attr in self.attributes:
            if my_attr.same_name(attr):
                return my_attr.value
        else:
            return ''


class Table(object):

    def __init__(self, name, datas, key=0):
        self.name = name
        self.recodes = []
        self.attributes = []
        self.keys_list = []


        if isinstance(key, int):
            key_idx = key
        elif isinstance(key, str):
            key_idx = datas[0].index(key)

        self.attributes = [Attribute(attr_name) for attr_name in datas[0]]
        self.key = self.attributes[key_idx]

        for idx, rec_data in enumerate(datas[1:]):
            if len(rec_data) != len(self.attributes):
                raise ValueError(f' not match the number of {idx=} recode attributes ({len(rec_data)}) and the number of table attributes ({len(self.attributes)}) ')

            recode_name = rec_data[key_idx]
            if not recode_name:
                raise ValueError(f'\n\n\t the key of line {idx} recode is None. \n')
            if self.has_recode(recode_name):
                raise SignalExistError(f'{recode_name} Recode already exists.')
            attributes = [attr.set(data) for attr, data in zip(self.attributes, rec_data)]
            self.recodes.append(Recode(key_idx, attributes))
            self.keys_list.append(attributes[key_idx])


    def __repr__(self):
        return f'Table(\'{self.name}\')'

    def get_attr(self, attr):
        if not attr in self.attributes:
            return UndifinedAttribute('Undifined')

        return self.attributes[self.attributes.index(attr)]

    def get_recode(self, recode_name):
        if not self.has_recode(recode_name):
            return None
        return self.recodes[self.recode_name_list.index(recode_name)]

    def get_field(self, key, attr):
        attr_idx = self.attributes.index(attr)
        recode = self.recodes[self.keys_list.index(key)]
        return recode.attributes[attr_idx]

    def has_attr(self, attr):
        return attr in self.attributes

    def has_recode(self, recode):
        if isinstance(recode, str):
            return recode in [rec.get_value(rec.key) for rec in self.recodes]
        elif isinstance(recode, Attribute):
            return recode in self.keys_list
        else:
            # todo: 主キー名、主キー以外を渡されたときの対処
            raise ValueError()


class ExcelTool(object):

    def __init__(self, path, sheet_name=None, anchor=(2,2)):
        self.path = Path(path)
        self.name = sheet_name
        self.anchor = anchor

        if not self.path.exists():
            wb = px.Workbook()
            font = px.styles.Font(name='Meiryo UI')
            for row in wb.worksheets[0].iter_rows():
                for cell in row:
                    cell.fornt = font
            wb.save(path)


        try:
            wb = px.load_workbook(path)
            self.wb = wb
        except PermissionError:
            raise PermissionError(f'\n\n\t please  close the {path}. \n') 

        if sheet_name is None:
            ws = wb.worksheets[0]
        else:
            ws = wb[sheet_name]
        self.ws = ws

    def clear_more_than(self, col=None, row=None):
        col += self.anchor[0]
        row += self.anchor[1]
        no_borders = px.styles.borders.Border(
            left=None, top=None, right=None, bottom=None
        )
        if not col is None:
            for row_idx in range(1, self.ws.max_row+1):
                cell = self.ws.cell(row=row_idx, column=col)
                border = px.styles.Border(left=cell.border.left)
                cell.border = border
                cell.fill = px.styles.PatternFill(fill_type=None)

            for col_idx in range(col + 1, self.ws.max_column+1):
                for row_idx in range(1, self.ws.max_row+1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    print(f'del border {cell.coordinate}')
                    cell.border = no_borders
                    cell.fill = px.styles.PatternFill(fill_type=None)

        if not row is None:
            for col_idx in range(1, self.ws.max_column+1):
                cell = self.ws.cell(row=row, column=col_idx)
                border = px.styles.Border(top=cell.border.top)
                cell.border = border
                cell.fill = px.styles.PatternFill(fill_type=None)

            for row_idx in range(row + 1, self.ws.max_row+1):
                for row_idx in range(1, self.ws.max_column+1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    cell.border = no_borders
                    cell.fill = px.styles.PatternFill(fill_type=None)
        self.save()


    def to_list(self, min_row=None, max_row=None, min_col=None, max_col=None):
        """セルの値をリストにして出力。値の先頭末尾の空白は削除。空白の場合は空文字。"""
        min_row = self.ws.min_row if min_row is None else min_row
        min_col = self.ws.min_column if min_col is None else min_col
        max_row = self.ws.max_row if max_row is None else max_row
        max_col = self.ws.max_column if max_col is None else max_col

        out_list = []
        for r in self.ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col, values_only=True):
            out_list.append(list(map(lambda x: str(x).strip() if not x is None else '', r)))

        return out_list

    def fill(self, value=None, color=None, row=None, min_row=None, min_col=None, max_row=None, max_col=None):

        min_row = self.anchor[1] if min_row is None else min_row
        min_col = self.anchor[0] if min_col is None else min_col
        max_row = self.ws.max_row if max_row is None else max_row
        max_col = self.ws.max_column if max_col is None else max_col

        if not value is None:
            for row in self.ws.iter_rows(min_row=min_row, max_row=max_row):
                for cell in row:
                    if cell.value == value:
                        cell.fill = px.styles.PatternFill(fill_type='solid', start_color=color)
            self.save()
            return
        if not row is None:
            fill = px.styles.PatternFill(fill_type='solid', start_color=color)
            for row_idx in [row + self.anchor[1] for row in row]:
                for col_idx in range(min_col, max_col+1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill

            self.save()
            return

    def line(self, cols=None, rows=None, type='thin'):
        if not rows is None:
            for row_idx in rows:
                row_idx += self.anchor[1]
                for col_idx in range(self.anchor[0], self.ws.max_column+1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    exist_border = cell.border
                    border = px.styles.Border(left=exist_border.left,
                                              top=px.styles.Side(style=type),
                                              right=exist_border.right,
                                              bottom=exist_border.bottom)
                    cell.border = border

        if not cols is None:
            for col_idx in cols:
                col_idx += self.anchor[1]
                for row_idx in range(self.anchor[1], self.ws.max_row+1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    exist_border = cell.border
                    border = px.styles.Border(left=px.styles.Side(style=type),
                                              top=exist_border.top,
                                              right=exist_border.right,
                                              bottom=exist_border.bottom)
                    cell.border = border

        self.save()

    def line_thin(self, cols, rows):
        self.line(cols, rows, type='thin')

    def lie_dotted(self, cols, rows):
        self.line(cols, rows, 'dotted')

    def write(self, data_list):
        for row_i, row_data in enumerate(data_list, start=self.anchor[1]):
            for col_i, data in enumerate(row_data, start=self.anchor[0]):
                self.ws.cell(row=row_i, column=col_i, value=data)
        self.save()
        print(f'write path "{self.path}"')

    def save(self):
        self.wb.save(str(self.path))

if __name__ == "__main__":

    data_path1 = 'diff-tools/sample/Signals_ver1.xlsx'
    data_path2 = 'diff-tools/sample/Signals_ver2.xlsx'
    diff_path = 'diff-tools/sample/Signals_diff.xlsx'

    datas1 = ExcelTool(data_path1).to_list()
    datas2 = ExcelTool(data_path2).to_list(max_row=13)
    t1 = Table(1, datas1)
    t2 = Table(2, datas2)

    all_tables = [t1, t2]

    # get all attributes
    comparison_attributes = []
    for t in all_tables:
        for attr in t.attributes:
            if attr == t.key:
                continue
            if not attr in comparison_attributes:
                comparison_attributes.append(attr)

    # get all recodes name
    all_keys = []
    for t in all_tables:
        for key in t.keys_list:
            if not key in all_keys:
                all_keys.append(key)

    diff_datas = [[attr.name for attr in [all_tables[0].key] + comparison_attributes]]
    for key in all_keys:
        diff_data = [key.value]
        for attr in comparison_attributes:
            result = Judgement.NOJUDGEMENT
            for i in range(len(all_tables) - 1):
                t1 = all_tables[i]
                t2 = all_tables[i+1]
                # いずれかのテーブルで、判定したい属性が未定義ならその属性の判定は判定無しとする。
                if not t1.has_attr(attr) or not t2.has_attr(attr):
                    result = Judgement.NOJUDGEMENT
                    continue

                # まだ判定がされていない場合
                if result == Judgement.NOJUDGEMENT:
                    if t1.has_recode(key) and not t2.has_recode(key):
                        result = Judgement.DEL
                    elif not t1.has_recode(key) and t2.has_recode(key):
                        result = Judgement.ADD
                    elif t1.get_field(key, attr) == t2.get_field(key, attr):
                        result = Judgement.SAME
                    else:
                        result = Judgement.CHANGE
                # すでに判定されている（テーブルが3以上ある）場合
                else:
                    if result == Judgement.SAME and t1.get_field(key, attr) == t2.get_field(key, attr):
                        result = Judgement.SAME
                    else:
                        result = Judgement.CHANGE
            diff_data.append(result.name)
        diff_datas.append(diff_data)

    all_tables.append(Table('judgement', diff_datas))


    all_datas = []
    all_attributes = [all_tables[0].key.name]
    all_attributes.extend([f'{t.name}: {t.get_attr(attr).name}' for attr in comparison_attributes for t in all_tables])
    all_datas.append(all_attributes)


    for key in all_keys:
        data = [key.value]
        for comp_attr in comparison_attributes:
            for t in all_tables:
                if not t.has_attr(comp_attr):
                    data.append('')
                    continue
                if not t.has_recode(key):
                    data.append('-')
                    continue

                data.append(t.get_field(key, comp_attr).value)
        all_datas.append(data)


    out = ExcelTool(diff_path)
    out.write(all_datas)

    out.fill(value=Judgement.SAME.name, color='BDD7EE')
    out.fill(value=Judgement.ADD.name, color='C6E0B4')
    out.fill(value=Judgement.DEL.name, color='DBDBDB')
    out.fill(value=Judgement.CHANGE.name, color='FFE699')
    out.fill(value=Judgement.NOJUDGEMENT.name, color='C6ACD9')
    out.fill(row=[0], color='F8CBAD')

    thin_line_col = [0]
    thin_line_col.extend([c for c in range(1, len(all_tables)*(len(comparison_attributes)+1), len(all_tables))])

    dot_line_col = []
    for c in range(len(comparison_attributes)):
        for j in range(len(all_tables) - 1):
            dot_line_col.append(2 + c*len(all_tables) + j)

    thin_line_row = [0, 1, len(all_keys)+1]
    dot_line_row = [r for r in range(2, len(all_keys)+1)]

    out.line(cols=thin_line_col, type='thin')
    out.line(cols=dot_line_col, type='dotted')
    out.line(rows=thin_line_row, type='thin')
    out.line(rows=dot_line_row, type='dotted')

    out.clear_more_than(len(all_attributes), len(all_keys) + 1)
