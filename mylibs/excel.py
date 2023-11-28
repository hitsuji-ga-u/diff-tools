from pathlib import Path
import openpyxl as px


class ExcelHelper(object):

    def __init__(self, path, sheet_name=None, anchor=(2,2)):
        self.path = Path(path)
        self.name = sheet_name
        self.anchor = anchor

        if not self.path.exists():
            wb = px.Workbook()
            wb.save(path)

        try:
            wb = px.load_workbook(path)
            self.wb = wb
        except PermissionError:
            raise PermissionError(f'\n\n\t please  close the {path}. \n') 

        if sheet_name is None:
            ws = wb.worksheets[0]
        else:
            ws = wb[sheet_name]
        self.ws = ws

    def clear_more_than(self, col=None, row=None):
        col += self.anchor[0]
        row += self.anchor[1]
        no_borders = px.styles.borders.Border(
            left=None, top=None, right=None, bottom=None
        )
        if not col is None:
            for row_idx in range(1, self.ws.max_row+1):
                cell = self.ws.cell(row=row_idx, column=col)
                border = px.styles.Border(left=cell.border.left)
                cell.border = border
                cell.fill = px.styles.PatternFill(fill_type=None)

            for col_idx in range(col + 1, self.ws.max_column+1):
                for row_idx in range(1, self.ws.max_row+1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    cell.border = no_borders
                    cell.fill = px.styles.PatternFill(fill_type=None)

        if not row is None:
            for col_idx in range(1, self.ws.max_column+1):
                cell = self.ws.cell(row=row, column=col_idx)
                border = px.styles.Border(top=cell.border.top)
                cell.border = border
                cell.fill = px.styles.PatternFill(fill_type=None)

            for row_idx in range(row + 1, self.ws.max_row+1):
                for row_idx in range(1, self.ws.max_column+1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    cell.border = no_borders
                    cell.fill = px.styles.PatternFill(fill_type=None)
        self.save()


    def to_list(self, min_row=None, max_row=None, min_col=None, max_col=None):
        """セルの値をリストにして出力。値の先頭末尾の空白は削除。空白の場合は空文字。"""
        min_row = self.ws.min_row if min_row is None else min_row
        min_col = self.ws.min_column if min_col is None else min_col
        max_row = self.ws.max_row if max_row is None else max_row
        max_col = self.ws.max_column if max_col is None else max_col

        out_list = []
        for r in self.ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col, values_only=True):
            out_list.append(list(map(lambda x: str(x).strip() if not x is None else '', r)))

        return out_list

    def fill(self, value=None, color=None, row=None, min_row=None, min_col=None, max_row=None, max_col=None):

        min_row = self.anchor[1] if min_row is None else min_row
        min_col = self.anchor[0] if min_col is None else min_col
        max_row = self.ws.max_row if max_row is None else max_row
        max_col = self.ws.max_column if max_col is None else max_col
        if not value is None:
            for row in self.ws.iter_rows(min_row=min_row, max_row=max_row):
                for cell in row:
                    if cell.value == value:
                        cell.fill = px.styles.PatternFill(fill_type='solid', start_color=color)
            self.save()
            return
        if not row is None:
            fill = px.styles.PatternFill(fill_type='solid', start_color=color)
            for row_idx in [row + self.anchor[1] for row in row]:
                for col_idx in range(min_col, max_col+1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill

            self.save()
            return

    def font(self, font):
        f = px.styles.Font(name=font)
        for row in range(1, self.ws.max_row+1):
            for col in range(1, self.ws.max_column+1):
                self.ws.cell(row=row, column=col).font = f
        self.save()

    def line(self, cols=None, rows=None, type='thin'):
        if not rows is None:
            for row_idx in rows:
                row_idx += self.anchor[1]
                for col_idx in range(self.anchor[0], self.ws.max_column+1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    exist_border = cell.border
                    border = px.styles.Border(left=exist_border.left,
                                              top=px.styles.Side(style=type),
                                              right=exist_border.right,
                                              bottom=exist_border.bottom)
                    cell.border = border

        if not cols is None:
            for col_idx in cols:
                col_idx += self.anchor[1]
                for row_idx in range(self.anchor[1], self.ws.max_row+1):
                    cell = self.ws.cell(row=row_idx, column=col_idx)
                    exist_border = cell.border
                    border = px.styles.Border(left=px.styles.Side(style=type),
                                              top=exist_border.top,
                                              right=exist_border.right,
                                              bottom=exist_border.bottom)
                    cell.border = border

        self.save()

    def line_thin(self, cols, rows):
        self.line(cols, rows, type='thin')

    def lie_dotted(self, cols, rows):
        self.line(cols, rows, 'dotted')

    def write(self, data_list):
        for row_i, row_data in enumerate(data_list, start=self.anchor[1]):
            for col_i, data in enumerate(row_data, start=self.anchor[0]):
                self.ws.cell(row=row_i, column=col_i, value=data)
        self.save()
        print(f'write path "{self.path}"')

    def save(self):
        self.wb.save(str(self.path))

